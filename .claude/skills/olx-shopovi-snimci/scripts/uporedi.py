"""Poredjenje dva snimka PIK/OLX shopova.

Utvrdjuje sta se promijenilo izmedju starijeg i novijeg snimka: novi shopovi,
nestali shopovi, promjene paketa, promjene broja oglasa, i shopovi koji su
presli iz praznog u aktivno stanje ili obrnuto.

Oba fajla se samo citaju.

Upotreba:
    python3 uporedi.py <stari.xlsx> <novi.xlsx> [izlaz.xlsx]

Bez zadatog izlaza pravi <folder novog>/shopovi-razlika-<datum starog>-do-<datum novog>.xlsx
"""
import re
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

MASTER = "SVI - zbirno"
KLJUC = "Shop (username)"
KOL_PAKET, KOL_OGLASI = "PIK paket", "Broj oglasa"
KOL_NAZIV, KOL_KANTON, KOL_GRAD = "Puni naziv / Firma", "Kanton / Regija", "Grad / Opština"
KOL_LINK, KOL_REG = "Link", "Registrovan"


def datum_iz_imena(p: Path) -> str:
    m = re.search(r"(\d{4}-\d{2}-\d{2})", p.name)
    return m.group(1) if m else "nepoznat"


def ucitaj(p: Path) -> pd.DataFrame:
    listovi = pd.ExcelFile(p, engine="openpyxl").sheet_names
    if MASTER not in listovi:
        sys.exit(f"GRESKA: nema lista {MASTER!r} u {p.name}. Postojeci: {listovi}")
    df = pd.read_excel(p, sheet_name=MASTER, engine="openpyxl")
    for k in (KLJUC, KOL_PAKET, KOL_OGLASI):
        if k not in df.columns:
            sys.exit(f"GRESKA: {p.name} nema kolonu {k!r}. Postojece: {list(df.columns)}")
    df[KOL_OGLASI] = pd.to_numeric(df[KOL_OGLASI], errors="coerce").fillna(0).astype(int)
    dup = int(df[KLJUC].duplicated().sum())
    if dup:
        print(f"  UPOZORENJE: {p.name} ima {dup} duplih usernamea. Uzima se prvi red po usernameu.")
        df = df.drop_duplicates(subset=[KLJUC], keep="first")
    return df


def formatiraj(dst: Path) -> None:
    wb = load_workbook(dst)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
        zaglavlje = [c.value for c in ws[1]]
        for c in ws[1]:
            c.font = Font(bold=True)
        if KOL_LINK in zaglavlje:
            i = zaglavlje.index(KOL_LINK) + 1
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=i)
                if isinstance(c.value, str) and c.value.startswith("http"):
                    c.hyperlink = c.value
                    c.font = Font(color="0563C1", underline="single")
        for i, naslov in enumerate(zaglavlje, start=1):
            duzine = [len(str(naslov))]
            for r in range(2, min(ws.max_row, 300) + 1):
                v = ws.cell(row=r, column=i).value
                if v is not None:
                    duzine.append(len(str(v)))
            ws.column_dimensions[get_column_letter(i)].width = min(max(duzine) + 2, 50)
    wb.save(dst)


def main() -> None:
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    stari_p, novi_p = (Path(a).expanduser().resolve() for a in sys.argv[1:3])
    for p in (stari_p, novi_p):
        if not p.exists():
            sys.exit(f"GRESKA: nema fajla {p}")
    d_stari, d_novi = datum_iz_imena(stari_p), datum_iz_imena(novi_p)
    dst = Path(sys.argv[3]).expanduser().resolve() if len(sys.argv) > 3 \
        else novi_p.parent / f"shopovi-razlika-{d_stari}-do-{d_novi}.xlsx"

    print(f"Stari snimak: {stari_p.name} ({d_stari})")
    stari = ucitaj(stari_p)
    print(f"  {len(stari)} shopova")
    print(f"Novi snimak:  {novi_p.name} ({d_novi})")
    novi = ucitaj(novi_p)
    print(f"  {len(novi)} shopova")

    s_kljuc = set(stari[KLJUC])
    n_kljuc = set(novi[KLJUC])

    opis = [c for c in (KLJUC, KOL_NAZIV, KOL_PAKET, KOL_GRAD, KOL_KANTON,
                        KOL_OGLASI, KOL_REG, KOL_LINK) if c in novi.columns]
    opis_s = [c for c in opis if c in stari.columns]

    novi_shopovi = novi[novi[KLJUC].isin(n_kljuc - s_kljuc)][opis].sort_values(
        KOL_OGLASI, ascending=False)
    nestali = stari[stari[KLJUC].isin(s_kljuc - n_kljuc)][opis_s].sort_values(
        KOL_OGLASI, ascending=False)

    # Zajednicki shopovi: spoji stanje pa uporedi
    z = stari.merge(novi, on=KLJUC, suffixes=("_stari", "_novi"))
    z["razlika oglasa"] = z[f"{KOL_OGLASI}_novi"] - z[f"{KOL_OGLASI}_stari"]

    kol_paket = [KLJUC, f"{KOL_NAZIV}_novi", f"{KOL_PAKET}_stari", f"{KOL_PAKET}_novi",
                 f"{KOL_KANTON}_novi", f"{KOL_OGLASI}_novi"]
    kol_paket = [c for c in kol_paket if c in z.columns]
    paket_promjena = z[z[f"{KOL_PAKET}_stari"] != z[f"{KOL_PAKET}_novi"]][kol_paket]

    kol_ogl = [KLJUC, f"{KOL_NAZIV}_novi", f"{KOL_PAKET}_novi", f"{KOL_KANTON}_novi",
               f"{KOL_OGLASI}_stari", f"{KOL_OGLASI}_novi", "razlika oglasa"]
    kol_ogl = [c for c in kol_ogl if c in z.columns]
    ogl_promjena = z[z["razlika oglasa"] != 0][kol_ogl].sort_values(
        "razlika oglasa", ascending=False)

    ozivjeli = z[(z[f"{KOL_OGLASI}_stari"] == 0) & (z[f"{KOL_OGLASI}_novi"] > 0)][kol_ogl] \
        .sort_values(f"{KOL_OGLASI}_novi", ascending=False)
    ugasili = z[(z[f"{KOL_OGLASI}_stari"] > 0) & (z[f"{KOL_OGLASI}_novi"] == 0)][kol_ogl] \
        .sort_values(f"{KOL_OGLASI}_stari", ascending=False)

    sazetak = pd.DataFrame({
        "polje": [
            "Stari snimak", "Datum starog", "Novi snimak", "Datum novog",
            "Shopova u starom", "Shopova u novom", "Neto promjena",
            "Novih shopova", "Nestalih shopova", "Zajednickih",
            "Promijenili paket", "Promijenili broj oglasa",
            "Iz praznog u aktivno", "Iz aktivnog u prazno",
            "Ukupno oglasa stari", "Ukupno oglasa novi", "Razlika oglasa",
            "Generisano",
        ],
        "vrijednost": [
            stari_p.name, d_stari, novi_p.name, d_novi,
            len(stari), len(novi), len(novi) - len(stari),
            len(novi_shopovi), len(nestali), len(z),
            len(paket_promjena), len(ogl_promjena),
            len(ozivjeli), len(ugasili),
            int(stari[KOL_OGLASI].sum()), int(novi[KOL_OGLASI].sum()),
            int(novi[KOL_OGLASI].sum() - stari[KOL_OGLASI].sum()),
            date.today().isoformat(),
        ],
    })

    with pd.ExcelWriter(dst, engine="openpyxl") as w:
        sazetak.to_excel(w, sheet_name="Sazetak", index=False)
        novi_shopovi.to_excel(w, sheet_name="Novi shopovi", index=False)
        nestali.to_excel(w, sheet_name="Nestali shopovi", index=False)
        paket_promjena.to_excel(w, sheet_name="Promjena paketa", index=False)
        ozivjeli.to_excel(w, sheet_name="Iz praznog u aktivno", index=False)
        ugasili.to_excel(w, sheet_name="Iz aktivnog u prazno", index=False)
        ogl_promjena.to_excel(w, sheet_name="Promjena broja oglasa", index=False)

    formatiraj(dst)

    print("\n" + "=" * 58)
    print(f"RAZLIKA {d_stari} -> {d_novi}")
    print("=" * 58)
    print(f"Shopova: {len(stari)} -> {len(novi)} ({len(novi)-len(stari):+d})")
    print(f"Novih:                {len(novi_shopovi)}")
    print(f"Nestalih:             {len(nestali)}")
    print(f"Promijenili paket:    {len(paket_promjena)}")
    print(f"Iz praznog u aktivno: {len(ozivjeli)}")
    print(f"Iz aktivnog u prazno: {len(ugasili)}")
    print(f"Oglasa: {int(stari[KOL_OGLASI].sum())} -> {int(novi[KOL_OGLASI].sum())} "
          f"({int(novi[KOL_OGLASI].sum()-stari[KOL_OGLASI].sum()):+d})")
    print(f"\nSnimljeno: {dst}")


if __name__ == "__main__":
    main()
