"""Provjera da li shop iz prodajne liste zaista izdvaja oglase i trosi kredite.

Cita javni endpoint GET /users/:username/listings, koji radi i bez tokena, pa
ovaj skript ne dira nijedan nalog i ne trosi nijedan kredit. Samo cita.

Nalaz iz probe 26.07.2026.: izdvojeni oglasi dolaze prvi u nizu, pa je dovoljno
prelistavati do prve stranice bez izdvojenog oglasa. Pretpostavka nije stroga
(vidjen je jedan neizdvojen oglas unutar prve stranice), zato skript ide jednu
stranicu dalje od prve ciste, a sa --validiraj N provjerava stopu promasaja
punim skeniranjem uzorka shopova.

Sta se biljezi po shopu: broj aktivnih, broj izdvojenih i po tipu, akcijske
cijene, OLX stories, zadnja obnova i dani od nje, pa ocjena koliko shop stvarno
koristi platformu.

Upotreba:
    python3 izdvajanje_provjera.py <prodajna-lista.xlsx> [izlaz.xlsx]
        [--listovi "A prioritet,B prioritet"] [--pauza 0.35] [--validiraj 25]

Kes se pise u <izlaz bez ekstenzije>-kes.json, pa se prekinut posao nastavlja.
"""
import json
import random
import re
import sys
import time
import urllib.error
import urllib.request
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = "https://api.olx.ba"
PER_PAGE = 100
MAX_STRANICA = 6          # gornja granica prelistavanja po shopu
POMOCNI = {"Sazetak", "Sažetak", "Info", "Analiza"}
KOL_SHOP = "Shop (username)"


def get(path: str, pauza: float, pokusaj: int = 0):
    r = urllib.request.Request(BASE + path, headers={
        "Accept": "application/json",
        "User-Agent": "olx-toolkit-analiza/1.0",
    })
    try:
        with urllib.request.urlopen(r, timeout=30) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        if e.code in (429, 500, 502, 503) and pokusaj < 4:
            time.sleep(min(2 ** pokusaj * 2, 30))
            return get(path, pauza, pokusaj + 1)
        raise
    finally:
        time.sleep(pauza)


def skeniraj(username: str, pauza: float, sve_stranice: bool = False) -> dict:
    """Vrati mjere za jedan shop. sve_stranice sluzi samo za validaciju."""
    izdvojeni = 0
    po_tipu: dict = {}
    akcije = stories = pinned = 0
    zadnja = 0
    procitano = 0
    stranica = 1
    ciste = 0
    total = None
    while stranica <= (MAX_STRANICA if not sve_stranice else 999):
        d = get(f"/users/{username}/listings?page={stranica}&per_page={PER_PAGE}", pauza)
        meta = d.get("meta") or {}
        total = meta.get("total", total)
        red = d.get("data") or []
        if not red:
            break
        procitano += len(red)
        na_str = 0
        for x in red:
            s = x.get("sponsored") or 0
            if s:
                izdvojeni += 1
                na_str += 1
                po_tipu[str(s)] = po_tipu.get(str(s), 0) + 1
            if x.get("has_discount"):
                akcije += 1
            if x.get("olx_stories"):
                stories += 1
            if x.get("pinned"):
                pinned += 1
            dt = x.get("date") or 0
            if isinstance(dt, (int, float)) and dt > zadnja:
                zadnja = int(dt)
        if not sve_stranice:
            ciste = ciste + 1 if na_str == 0 else 0
            # Jedna cista stranica moze biti slucajna, dvije znace da je kraj bloka.
            if ciste >= 2:
                break
        if stranica >= (meta.get("last_page") or 1):
            break
        stranica += 1
    return {
        "aktivnih": total if total is not None else procitano,
        "procitano": procitano,
        "izdvojenih": izdvojeni,
        "po_tipu": po_tipu,
        "akcijske_cijene": akcije,
        "olx_stories": stories,
        "pinned": pinned,
        "zadnja_obnova_ts": zadnja,
        "stranica": stranica,
    }


def klasa(izdv: int, aktivnih: int) -> str:
    if aktivnih == 0:
        return "nema oglasa"
    if izdv == 0:
        return "ne izdvaja"
    dio = izdv / aktivnih * 100
    if izdv >= 20 or dio >= 10:
        return "izdvaja jako"
    if izdv >= 5:
        return "izdvaja redovno"
    return "izdvaja malo"


def formatiraj(dst: Path) -> None:
    zeleno = PatternFill("solid", fgColor="E2EFDA")
    crveno = PatternFill("solid", fgColor="FCE4E4")
    wb = load_workbook(dst)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        zag = [c.value for c in ws[1]]
        for c in ws[1]:
            c.font = Font(bold=True)
        if "klasa koristenja" in zag:
            i = zag.index("klasa koristenja") + 1
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=i).value
                if v in ("izdvaja jako", "izdvaja redovno"):
                    ws.cell(row=r, column=i).fill = zeleno
                elif v == "ne izdvaja":
                    ws.cell(row=r, column=i).fill = crveno
        if "Link" in zag:
            i = zag.index("Link") + 1
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=i)
                if isinstance(c.value, str) and c.value.startswith("http"):
                    c.hyperlink = c.value
                    c.font = Font(color="0563C1", underline="single")
        for i, naslov in enumerate(zag, start=1):
            duz = [len(str(naslov))]
            for r in range(2, min(ws.max_row, 300) + 1):
                v = ws.cell(row=r, column=i).value
                if v is not None:
                    duz.append(len(str(v)))
            ws.column_dimensions[get_column_letter(i)].width = min(max(duz) + 2, 46)
    wb.save(dst)


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    # Vrijednost opcije ne smije zavrsiti kao pozicioni argument, zato se
    # indeksi potroseni na opcije preskacu.
    opcije = {}
    potroseni = set()
    for i, a in enumerate(sys.argv[1:], start=1):
        if a.startswith("--") and i + 1 < len(sys.argv):
            opcije[a[2:]] = sys.argv[i + 1]
            potroseni.update({i, i + 1})
    pozicioni = [a for i, a in enumerate(sys.argv[1:], start=1)
                 if i not in potroseni and not a.startswith("--")]
    src = Path(pozicioni[0]).expanduser().resolve()
    m = re.search(r"(\d{4}-\d{2}-\d{2})", src.name)
    datum = m.group(1) if m else date.today().isoformat()
    dst = Path(pozicioni[1]).expanduser().resolve() if len(pozicioni) > 1 \
        else src.parent / f"izdvajanje-provjera-{datum}.xlsx"
    kes_put = dst.with_name(dst.stem + "-kes.json")
    pauza = float(opcije.get("pauza", 0.35))
    validiraj = int(opcije.get("validiraj", 0))

    xl = pd.ExcelFile(src, engine="openpyxl")
    listovi = [s.strip() for s in opcije["listovi"].split(",")] if "listovi" in opcije \
        else [s for s in xl.sheet_names if s not in POMOCNI]
    dijelovi = []
    for s in listovi:
        d = pd.read_excel(src, sheet_name=s, engine="openpyxl")
        d["izvorni list"] = s
        dijelovi.append(d)
    df = pd.concat(dijelovi, ignore_index=True).drop_duplicates(subset=[KOL_SHOP])
    print(f"Ulaz: {src.name}, listovi {listovi}, shopova {len(df)}")

    kes = json.loads(kes_put.read_text()) if kes_put.exists() else {}
    print(f"Kes: {len(kes)} shopova od prije")
    greske = 0
    for n, u in enumerate(df[KOL_SHOP].astype(str), start=1):
        if u in kes:
            continue
        try:
            kes[u] = skeniraj(u, pauza)
        except Exception as e:            # nedostupan shop ne smije zaustaviti posao
            kes[u] = {"greska": str(e)[:120]}
            greske += 1
        if n % 25 == 0:
            kes_put.write_text(json.dumps(kes, ensure_ascii=False))
            print(f"  {n}/{len(df)} obradjeno, gresaka {greske}", flush=True)
    kes_put.write_text(json.dumps(kes, ensure_ascii=False))

    danas = datetime.now()
    redovi = []
    for _, r in df.iterrows():
        u = str(r[KOL_SHOP])
        k = kes.get(u, {})
        akt = int(k.get("aktivnih") or 0)
        izdv = int(k.get("izdvojenih") or 0)
        ts = int(k.get("zadnja_obnova_ts") or 0)
        dana = (danas - datetime.fromtimestamp(ts)).days if ts else None
        redovi.append({
            KOL_SHOP: u,
            "Puni naziv / Firma": r.get("Puni naziv / Firma"),
            "PIK paket": r.get("PIK paket"),
            "prioritet": r.get("prioritet"),
            "izvorni list": r.get("izvorni list"),
            "Kanton / Regija": r.get("Kanton / Regija"),
            "aktivnih sada": akt,
            "procitano oglasa": int(k.get("procitano") or 0),
            "izdvojenih": izdv,
            "procenat izdvojenih": round(izdv / akt * 100, 1) if akt else 0.0,
            # Tipovi izdvajanja po MCP alatu: 1 klasicno, 2 premium (server.ts:645).
            "tip 1 klasicno": int((k.get("po_tipu") or {}).get("1", 0)),
            "tip 2 premium": int((k.get("po_tipu") or {}).get("2", 0)),
            "akcijske cijene": int(k.get("akcijske_cijene") or 0),
            "olx stories": int(k.get("olx_stories") or 0),
            "zadnja obnova": datetime.fromtimestamp(ts).date().isoformat() if ts else "",
            "dana od obnove": dana,
            "klasa koristenja": "greska" if k.get("greska") else klasa(izdv, akt),
            "napomena": k.get("greska", ""),
            "Link": r.get("Link"),
        })
    rez = pd.DataFrame(redovi)

    # Validacija pretpostavke da izdvojeni dolaze prvi: puni prolaz kroz uzorak.
    val = []
    if validiraj:
        kandidati = rez[(rez["izdvojenih"] == 0) & rez["aktivnih sada"].between(100, 800)]
        uzorak = kandidati[KOL_SHOP].tolist()
        random.seed(7)
        random.shuffle(uzorak)
        for u in uzorak[:validiraj]:
            puni = skeniraj(u, pauza, sve_stranice=True)
            val.append({KOL_SHOP: u, "aktivnih": puni["aktivnih"],
                        "procitano cijelo": puni["procitano"],
                        "izdvojenih u punom prolazu": puni["izdvojenih"]})
        print(f"Validacija: {len(val)} shopova, promasaja "
              f"{sum(1 for v in val if v['izdvojenih u punom prolazu'] > 0)}")

    grupe = rez.groupby("klasa koristenja").agg(
        shopova=(KOL_SHOP, "count"),
        Platinum=("PIK paket", lambda s: int((s == "Platinum").sum())),
        ukupno_izdvojenih=("izdvojenih", "sum"),
    ).reset_index().sort_values("shopova", ascending=False)

    with pd.ExcelWriter(dst, engine="openpyxl") as w:
        grupe.to_excel(w, sheet_name="Sazetak", index=False)
        rez.sort_values(["izdvojenih", "aktivnih sada"], ascending=False).to_excel(
            w, sheet_name="Svi shopovi", index=False)
        rez[rez["klasa koristenja"].isin(["izdvaja jako", "izdvaja redovno"])].sort_values(
            "izdvojenih", ascending=False).to_excel(w, sheet_name="Dokazano trose", index=False)
        rez[rez["klasa koristenja"] == "ne izdvaja"].sort_values(
            "aktivnih sada", ascending=False).to_excel(w, sheet_name="Ne izdvajaju", index=False)
        if val:
            pd.DataFrame(val).to_excel(w, sheet_name="Validacija", index=False)

    formatiraj(dst)
    print(grupe.to_string(index=False))
    print(f"\nSnimljeno: {dst}")


if __name__ == "__main__":
    main()
