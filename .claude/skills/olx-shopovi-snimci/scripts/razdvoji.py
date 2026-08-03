"""Razdvajanje PIK/OLX shopova po kantonima, sa oglasima i bez oglasa.

Ulaz je snimak shopova (xlsx sa listom "SVI - zbirno").
Izlaz je novi xlsx: po jedan list za svaki kanton, unutar lista prvo shopovi sa
oglasima pa prazni, a u svakom bloku Platinum prije Golda.

Originalni fajl se samo cita, nikad ne mijenja.

Upotreba:
    python3 razdvoji.py <snimak.xlsx> [izlaz.xlsx] [--paketi Bronze,Silver]

Bez zadatog izlaza pravi <folder snimka>/shopovi-razdvojeno-<datum iz imena>.xlsx

Sa --paketi izlaz sadrzi samo te pakete, dobija vlastite kolone po paketu u izvjestajima i
sufiks u imenu (npr. shopovi-razdvojeno-bronze-silver-<datum>.xlsx). Jedan poziv pravi jedan
fajl, pa se za dvije grupe skripta pokrece dvaput.
"""
import re
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

MASTER = "SVI - zbirno"

KOL = {
    "shop": "Shop (username)",
    "naziv": "Puni naziv / Firma",
    "paket": "PIK paket",
    "grad": "Grad / Opština",
    "kanton": "Kanton / Regija",
    "entitet": "Entitet",
    "oglasi": "Broj oglasa",
    "web": "Web stranica",
    "partner": "OLX partner",
    "reg": "Registrovan",
    "link": "Link",
}

# Redoslijed paketa pri sortiranju. Sve izvan ovoga ide na kraj bloka.
PAKET_RANG = {"Platinum": 0, "Gold": 1, "Silver": 2, "Bronze": 3}

# Paketi koji dobijaju vlastite kolone u izvjestaju kad --paketi nije zadat. Drzi se
# istorijskog izbora (Gold i Platinum), da default izlaz ostane isti kao dosad.
DEFAULT_PAKETI = ["Gold", "Platinum"]

RE_FIRMA_DOSLOVNO = re.compile(r"d\.o\.o\.?|s\.p\.?|d\.d\.?", re.IGNORECASE)
# Kratke oznake se traze kao samostalne rijeci, da "sp" ne pokupi Sport/Spektar.
RE_FIRMA_RIJEC = re.compile(r"(?<![\wčćžšđ])(doo|sp|dd|obrt)(?![\wčćžšđ])", re.IGNORECASE)
RE_DOMEN = re.compile(r"\.[A-Za-z]{2,}(?![A-Za-z])")
SMECE_PODNIZ = ("test", "qwerty", "asdf", "PUNI NAZIV", "Nezanam")
SMECE_TACNO = {"olx", "olx.ba", "pik.ba", "sve", "prodaja", "razno"}
NEDOZVOLJENO_U_IMENU = re.compile(r"[:\\/?*\[\]]")


def datum_iz_imena(putanja: Path) -> str:
    m = re.search(r"(\d{4}-\d{2}-\d{2})", putanja.name)
    return m.group(1) if m else date.today().isoformat()


def velicina(n: int) -> str:
    if n == 0:
        return "nula"
    if n <= 9:
        return "mali"
    if n <= 99:
        return "srednji"
    if n <= 999:
        return "veliki"
    return "vrlo veliki"


def je_firma(naziv, web) -> bool:
    n = "" if pd.isna(naziv) else str(naziv)
    if RE_FIRMA_DOSLOVNO.search(n) or RE_FIRMA_RIJEC.search(n):
        return True
    w = "" if pd.isna(web) else str(web).strip()
    return bool(w) and bool(RE_DOMEN.search(w))


def je_smece(naziv) -> bool:
    n = "" if pd.isna(naziv) else str(naziv).strip()
    if n.lower() in SMECE_TACNO:
        return True
    return any(p.lower() in n.lower() for p in SMECE_PODNIZ)


def ucitaj(src: Path) -> pd.DataFrame:
    listovi = pd.ExcelFile(src, engine="openpyxl").sheet_names
    if MASTER not in listovi:
        sys.exit(f"GRESKA: nema lista {MASTER!r} u {src.name}. Postojeci: {listovi}")
    df = pd.read_excel(src, sheet_name=MASTER, engine="openpyxl")
    nedostaju = [v for v in KOL.values() if v not in df.columns]
    if nedostaju:
        sys.exit(f"GRESKA: nedostaju kolone {nedostaju}\nPostojece: {list(df.columns)}")
    df[KOL["oglasi"]] = pd.to_numeric(df[KOL["oglasi"]], errors="coerce").fillna(0).astype(int)
    return df


def provjeri_zbir(src: Path, df: pd.DataFrame) -> None:
    listovi = pd.ExcelFile(src, engine="openpyxl").sheet_names
    pomocni = {MASTER, "Sažetak", "Sazetak"}
    po_lokaciji = [s for s in listovi if s not in pomocni]
    # Snimak moze imati samo zbirni list, bez listova po lokaciji. Tada nema sta da se
    # poredi, pa se to kaze, ne prijavljuje kao upozorenje o neslaganju.
    if not po_lokaciji:
        print("  Snimak ima samo zbirni list, nema listova po lokaciji za uporedbu.")
        return
    zbir = sum(len(pd.read_excel(src, sheet_name=s, engine="openpyxl")) for s in po_lokaciji)
    if zbir != len(df):
        print(f"  UPOZORENJE: zbir po listovima {zbir}, master {len(df)}, razlika {zbir-len(df):+d}")
    else:
        print(f"  Zbir po {len(po_lokaciji)} listova odgovara masteru ({zbir}).")


def oznake(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    o, p, n, w = KOL["oglasi"], KOL["paket"], KOL["naziv"], KOL["web"]
    df["ima_oglase"] = df[o].gt(0).map({True: "DA", False: "NE"})
    df["velicina"] = df[o].map(velicina)
    df["tip_naloga"] = "nejasno"
    df.loc[df[n].map(je_smece), "tip_naloga"] = "smece"
    df.loc[df.apply(lambda r: je_firma(r[n], r[w]), axis=1), "tip_naloga"] = "firma"
    df["godina_registracije"] = pd.to_datetime(df[KOL["reg"]], errors="coerce").dt.year
    df["_akt"] = df[o].eq(0).astype(int)
    df["_pak"] = df[p].map(PAKET_RANG).fillna(len(PAKET_RANG)).astype(int)
    return df


def pregled(df: pd.DataFrame, grupa: str, paketi: list,
            min_shopova: int = 0) -> pd.DataFrame:
    o, p = KOL["oglasi"], KOL["paket"]
    redovi = []
    for kljuc, g in df.groupby(grupa, dropna=False):
        if len(g) < min_shopova:
            continue
        bez = int(g[o].eq(0).sum())
        nenula = g.loc[g[o] > 0, o]
        akt = g[o].gt(0)
        red = {
            grupa: kljuc if pd.notna(kljuc) else "(prazno)",
            "broj shopova": len(g),
            "sa oglasima": len(g) - bez,
            "bez oglasa": bez,
            "procenat bez oglasa": round(bez / len(g) * 100, 1),
        }
        # Tri odvojene petlje, ne jedna: kolone su grupisane po pokazatelju (svi ukupni, pa
        # svi aktivni, pa svi bez oglasa), a ne po paketu. Jedna petlja bi tiho promijenila
        # redoslijed kolona onome ko ovaj list cita po poziciji.
        for pak in paketi:
            red[pak] = int(g[p].eq(pak).sum())
        for pak in paketi:
            red[f"{pak} aktivni"] = int((akt & g[p].eq(pak)).sum())
        for pak in paketi:
            red[f"{pak} bez oglasa"] = int((~akt & g[p].eq(pak)).sum())
        red["ukupno oglasa"] = int(g[o].sum())
        red["medijana oglasa (svi)"] = float(g[o].median())
        red["medijana oglasa (bez nula)"] = float(nenula.median()) if len(nenula) else 0.0
        redovi.append(red)
    return pd.DataFrame(redovi).sort_values("broj shopova", ascending=False)


def analiza(df: pd.DataFrame, pk: pd.DataFrame, datum: str, izvor: str,
            firmi_bez: int, paketi: list) -> tuple:
    """Prvi list izlaza: zbirno stanje i aktivni shopovi po paketu i kantonu.

    Sluzi kao radna podloga za pripremu liste profila za prodavce, pa su
    aktivni shopovi (najmanje jedan oglas) prva kolona i sortiranje.

    Imena kolona po paketu se grade istim obrascem kao u pregled(), jer se tabela po
    kantonima bira iz njenog rezultata PO IMENU.
    """
    o, p = KOL["oglasi"], KOL["paket"]
    akt = df[o].gt(0)
    ostali_akt = int((akt & ~df[p].isin(paketi)).sum())
    pokazatelj = ["Datum snimka", "Izvorni fajl", "Ukupno shopova",
                  "Aktivni (najmanje 1 oglas)"]
    vrijednost = [datum, izvor, len(df), int(akt.sum())]
    for pak in paketi:
        pokazatelj.append(f"{pak} aktivni")
        vrijednost.append(int((akt & df[p].eq(pak)).sum()))
    pokazatelj.append("Ostali paketi aktivni")
    vrijednost.append(ostali_akt)
    pokazatelj.append("Bez oglasa")
    vrijednost.append(int((~akt).sum()))
    for pak in paketi:
        pokazatelj.append(f"{pak} bez oglasa")
        vrijednost.append(int((~akt & df[p].eq(pak)).sum()))
    pokazatelj += ["Procenat bez oglasa", "Firme bez oglasa",
                   "Ukupno oglasa (aktivni)", "Medijana oglasa (aktivni)"]
    vrijednost += [f"{(~akt).mean()*100:.1f}%", firmi_bez,
                   int(df.loc[akt, o].sum()), float(df.loc[akt, o].median())]
    zbirno = pd.DataFrame({"pokazatelj": pokazatelj, "vrijednost": vrijednost})
    kol_kanton = pk.columns[0]
    kolone = ["sa oglasima"] + [f"{pak} aktivni" for pak in paketi] + ["bez oglasa"] \
        + [f"{pak} bez oglasa" for pak in paketi] \
        + ["procenat bez oglasa", "broj shopova", "ukupno oglasa",
           "medijana oglasa (bez nula)"]
    tabela = pk[[kol_kanton] + kolone].copy()
    tabela = tabela.rename(columns={
        "sa oglasima": "aktivni ukupno",
        "medijana oglasa (bez nula)": "medijana oglasa (aktivni)",
    }).sort_values("aktivni ukupno", ascending=False)
    zbir = {kol_kanton: "UKUPNO"}
    for c in tabela.columns[1:]:
        zbir[c] = int(tabela[c].sum()) if c not in (
            "procenat bez oglasa", "medijana oglasa (aktivni)") else ""
    zbir["procenat bez oglasa"] = round(
        int(tabela["bez oglasa"].sum()) / int(tabela["broj shopova"].sum()) * 100, 1)
    tabela = pd.concat([tabela, pd.DataFrame([zbir])], ignore_index=True)
    return zbirno, tabela


def formatiraj_analizu(ws, red_tabele: int, naslov_dodatak: str = "") -> None:
    """List Analiza ima dva bloka pa ne dobija filter ni zamrznut prvi red."""
    ws["A1"] = f"ANALIZA SNIMKA{naslov_dodatak} — aktivni shopovi po paketu i kantonu"
    ws["A1"].font = Font(bold=True, size=13)
    for red in (2, red_tabele):
        for c in ws[red]:
            c.font = Font(bold=True)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for i in range(1, ws.max_column + 1):
        duzine = [len(str(ws.cell(row=r, column=i).value or ""))
                  for r in range(2, ws.max_row + 1)]
        ws.column_dimensions[get_column_letter(i)].width = min(max(duzine) + 2, 42)


def formatiraj(dst: Path, red_tabele_analiza: int, naslov_dodatak: str = "") -> None:
    sivo = PatternFill("solid", fgColor="F2F2F2")
    wb = load_workbook(dst)
    if "Analiza" in wb.sheetnames:
        formatiraj_analizu(wb["Analiza"], red_tabele_analiza, naslov_dodatak)
    for ws in wb.worksheets:
        if ws.title == "Analiza":
            continue
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        zaglavlje = [c.value for c in ws[1]]
        for c in ws[1]:
            c.font = Font(bold=True)
        if KOL["link"] in zaglavlje:
            i = zaglavlje.index(KOL["link"]) + 1
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=i)
                if isinstance(c.value, str) and c.value.startswith("http"):
                    c.hyperlink = c.value
                    c.font = Font(color="0563C1", underline="single")
        if "ima_oglase" in zaglavlje:
            i = zaglavlje.index("ima_oglase") + 1
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=i).value == "NE":
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = sivo
        for i, naslov in enumerate(zaglavlje, start=1):
            duzine = [len(str(naslov))]
            for r in range(2, min(ws.max_row, 300) + 1):
                v = ws.cell(row=r, column=i).value
                if v is not None:
                    duzine.append(len(str(v)))
            ws.column_dimensions[get_column_letter(i)].width = min(max(duzine) + 2, 50)
    wb.save(dst)


def parsiraj_argumente(argv: list) -> tuple:
    """Rucno parsiranje, isti stil kao ostale skripte u ovom skillu (bez argparse).

    Vraca (pozicioni, trazeni_paketi); trazeni_paketi je None kad --paketi nije zadat.
    """
    pozicioni, trazeni = [], None
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == "--paketi":
            if i + 1 >= len(argv):
                sys.exit("GRESKA: --paketi trazi listu, npr. --paketi Bronze,Silver")
            trazeni = [s.strip() for s in argv[i + 1].split(",") if s.strip()]
            i += 2
            continue
        if a.startswith("--"):
            sys.exit(f"GRESKA: nepoznata opcija {a}")
        pozicioni.append(a)
        i += 1
    return pozicioni, trazeni


def uskladi_pakete(trazeni: list, prisutni) -> list:
    """Trazene pakete poredi sa onima koji STVARNO postoje u snimku, bez obzira na velicinu
    slova, i vraca ih u obliku kakav pise u snimku. Poredjenje ide protiv podatka, a ne
    protiv PAKET_RANG, da se ne odbije paket koji platforma uvede prije nego konstanta zna
    za njega."""
    mapa = {str(v).strip().lower(): str(v) for v in prisutni}
    izabrani, nepoznati = [], []
    for t in trazeni:
        kljuc = t.strip().lower()
        if kljuc not in mapa:
            nepoznati.append(t)
        elif mapa[kljuc] not in izabrani:
            izabrani.append(mapa[kljuc])
    if nepoznati:
        sys.exit(f"GRESKA: paket {nepoznati} ne postoji u snimku. "
                 f"Prisutni: {sorted(mapa.values())}")
    # Sortiranje po PAKET_RANG, ne po redoslijedu iz komande: isto ime fajla mora znaciti i
    # isti sadrzaj, a ovako se kolone poklapaju sa redoslijedom redova unutar kantonskog
    # lista. Paket koji PAKET_RANG ne zna ide na kraj, abecedno.
    return sorted(izabrani, key=lambda p: (PAKET_RANG.get(p, len(PAKET_RANG)), p))


def sufiks_paketa(paketi: list) -> str:
    """Abecedno i malim slovima, da ime fajla ne zavisi od redoslijeda u komandi:
    --paketi Bronze,Silver i --paketi Silver,Bronze daju isto bronze-silver."""
    return "-".join(sorted(p.lower() for p in paketi))


def main() -> None:
    pozicioni, trazeni = parsiraj_argumente(sys.argv[1:])
    if not pozicioni:
        sys.exit(__doc__)
    src = Path(pozicioni[0]).expanduser().resolve()
    if not src.exists():
        sys.exit(f"GRESKA: nema fajla {src}")
    datum = datum_iz_imena(src)

    print(f"Snimak: {src.name} (datum {datum})")
    df = ucitaj(src)
    print(f"  Master: {len(df)} redova, kolone se poklapaju.")
    provjeri_zbir(src, df)

    dup = df[KOL["shop"]].duplicated(keep=False)
    if dup.any():
        print(f"  UPOZORENJE: {int(dup.sum())} duplih redova po usernameu, "
              f"{df.loc[dup, KOL['shop']].nunique()} imena. Nijedan nije obrisan.")
    else:
        print("  Duplikata po usernameu nema.")

    df = oznake(df)
    van = sorted(set(df.loc[df["_pak"] >= len(PAKET_RANG), KOL["paket"]].dropna().astype(str)))
    if van:
        print(f"  NAPOMENA: paketi izvan poznatih idu na kraj bloka: {van}")
    poznati = df[KOL["paket"]].value_counts(dropna=False)
    print("  Paketi: " + ", ".join(f"{k} {v}" for k, v in poznati.items()))

    # Filter se primjenjuje POSLIJE oznake() i poslije dijagnostike iznad, da konzola uvijek
    # izvijesti o cijelom snimku, bez obzira koja grupa paketa ide u fajl.
    if trazeni is None:
        df_izlaz, paketi, naslov_dodatak, sufiks = df, DEFAULT_PAKETI, "", ""
    else:
        paketi = uskladi_pakete(trazeni, df[KOL["paket"]].dropna().unique())
        df_izlaz = df[df[KOL["paket"]].isin(paketi)]
        if df_izlaz.empty:
            sys.exit(f"GRESKA: nijedan shop nije u paketima {paketi}, fajl nije napravljen.")
        naslov_dodatak = f" ({', '.join(paketi)})"
        sufiks = f"-{sufiks_paketa(paketi)}"
        print(f"  Filter paketa: {', '.join(paketi)}, "
              f"{len(df_izlaz)} od {len(df)} shopova ide u izlaz.")

    dst = Path(pozicioni[1]).expanduser().resolve() if len(pozicioni) > 1 \
        else src.parent / f"shopovi-razdvojeno{sufiks}-{datum}.xlsx"

    pk = pregled(df_izlaz, KOL["kanton"], paketi)
    pg = pregled(df_izlaz, KOL["grad"], paketi, min_shopova=5)
    firme_bez = df_izlaz[df_izlaz[KOL["oglasi"]].eq(0)
                         & df_izlaz["tip_naloga"].eq("firma")].sort_values(
        [KOL["kanton"], KOL["grad"]], na_position="last")

    o = KOL["oglasi"]
    info = pd.DataFrame({
        "polje": ["Datum snimka", "Izvorni fajl", "Ukupno shopova", "Sa oglasima",
                  "Bez oglasa", "Procenat bez oglasa", "Firmi bez oglasa",
                  "Medijana oglasa (svi)", "Medijana oglasa (bez nula)", "Generisano",
                  "Paketi u izvjestaju"],
        "vrijednost": [
            datum, src.name, len(df_izlaz), int(df_izlaz[o].gt(0).sum()),
            int(df_izlaz[o].eq(0).sum()),
            f"{df_izlaz[o].eq(0).mean()*100:.1f}%", len(firme_bez),
            float(df_izlaz[o].median()),
            float(df_izlaz.loc[df_izlaz[o] > 0, o].median()),
            date.today().isoformat(),
            ", ".join(sorted(df_izlaz[KOL["paket"]].dropna().astype(str).unique())),
        ],
    })

    an_zbirno, an_tabela = analiza(df_izlaz, pk, datum, src.name, len(firme_bez), paketi)

    izlazne = [c for c in df_izlaz.columns if not c.startswith("_")]
    zauzeta = {"Analiza", "Info", "Pregled kantoni", "Pregled gradovi",
               "Firme bez oglasa"}

    def ime_lista(kanton: str) -> str:
        osnova = NEDOZVOLJENO_U_IMENU.sub("-", str(kanton))[:31].strip() or "Bez naziva"
        ime, i = osnova, 2
        while ime in zauzeta:
            suf = f" {i}"
            ime = osnova[: 31 - len(suf)] + suf
            i += 1
        zauzeta.add(ime)
        return ime

    ukupno = 0
    # Analiza je prvi list: zbirno od reda 2, tabela po kantonima ispod njega.
    red_tabele = 2 + len(an_zbirno) + 3
    with pd.ExcelWriter(dst, engine="openpyxl") as w:
        an_zbirno.to_excel(w, sheet_name="Analiza", index=False, startrow=1)
        an_tabela.to_excel(w, sheet_name="Analiza", index=False,
                           startrow=red_tabele - 1)
        info.to_excel(w, sheet_name="Info", index=False)
        pk.to_excel(w, sheet_name="Pregled kantoni", index=False)
        pg.to_excel(w, sheet_name="Pregled gradovi", index=False)
        for kanton in pk[KOL["kanton"]]:
            g = df_izlaz[df_izlaz[KOL["kanton"]].isna()] if kanton == "(prazno)" \
                else df_izlaz[df_izlaz[KOL["kanton"]] == kanton]
            g = g.sort_values(["_akt", "_pak", KOL["oglasi"]], ascending=[True, True, False])
            ukupno += len(g)
            g[izlazne].to_excel(w, sheet_name=ime_lista(kanton), index=False)
        firme_bez[izlazne].to_excel(w, sheet_name="Firme bez oglasa", index=False)

    formatiraj(dst, red_tabele, naslov_dodatak)
    print(f"  Zbir po kantonskim listovima: {ukupno} "
          f"({'odgovara' if ukupno == len(df_izlaz) else 'NE ODGOVARA'} izlazu)")
    print(f"\nSnimljeno: {dst}")


if __name__ == "__main__":
    main()
