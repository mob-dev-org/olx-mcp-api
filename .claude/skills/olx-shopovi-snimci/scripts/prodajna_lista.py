"""Prodajna lista potencijalnih klijenata iz snimka shopova.

Ulaz je razdvojeni izlaz skripta razdvoji.py (ima kolone ima_oglase, tip_naloga,
velicina, godina_registracije) ili sam snimak sa listom "SVI - zbirno".

Izlaz je novi xlsx sa prioritetima za prodavce: A, B i C po ocjeni, plus poseban
list praznih Platinum shopova koji placaju paket a nemaju ni jedan oglas.

Ocjena je heuristika, ne istina. Sluzi samo da prodavac zna koga zvati prvog.

Upotreba:
    python3 prodajna_lista.py <razdvojeno_ili_snimak.xlsx> [izlaz.xlsx] [--gold N --platinum N]
"""
import re
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

MASTER = "SVI - zbirno"
POMOCNI = {"Analiza", "Info", "Pregled kantoni", "Pregled gradovi",
           "Firme bez oglasa", "Sažetak", "Sazetak"}

KOL_SHOP = "Shop (username)"
KOL_NAZIV = "Puni naziv / Firma"
KOL_PAKET = "PIK paket"
KOL_GRAD = "Grad / Opština"
KOL_KANTON = "Kanton / Regija"
KOL_OGLASI = "Broj oglasa"
KOL_WEB = "Web stranica"
KOL_PARTNER = "OLX partner"
KOL_REG = "Registrovan"
KOL_LINK = "Link"

# Naknada po paketu, ulazi u procjenu vrijednosti pipelinea.
CIJENA = {"Gold": 50, "Platinum": 100, "Silver": 50, "Bronze": 50}

RE_FIRMA_DOSLOVNO = re.compile(r"d\.o\.o\.?|s\.p\.?|d\.d\.?", re.IGNORECASE)
RE_FIRMA_RIJEC = re.compile(r"(?<![\wčćžšđ])(doo|sp|dd|obrt)(?![\wčćžšđ])", re.IGNORECASE)
RE_DOMEN = re.compile(r"\.[A-Za-z]{2,}(?![A-Za-z])")


def ucitaj(src: Path) -> pd.DataFrame:
    listovi = pd.ExcelFile(src, engine="openpyxl").sheet_names
    if MASTER in listovi:
        df = pd.read_excel(src, sheet_name=MASTER, engine="openpyxl")
    else:
        dijelovi = [pd.read_excel(src, sheet_name=s, engine="openpyxl")
                    for s in listovi if s not in POMOCNI]
        if not dijelovi:
            sys.exit(f"GRESKA: u {src.name} nema ni {MASTER!r} ni kantonalnih listova.")
        df = pd.concat(dijelovi, ignore_index=True)
        df = df.drop_duplicates(subset=[KOL_SHOP])
    df[KOL_OGLASI] = pd.to_numeric(df[KOL_OGLASI], errors="coerce").fillna(0).astype(int)
    if "tip_naloga" not in df.columns:
        df["tip_naloga"] = df.apply(lambda r: "firma" if je_firma(r[KOL_NAZIV], r[KOL_WEB])
                                    else "nejasno", axis=1)
    if "godina_registracije" not in df.columns:
        df["godina_registracije"] = pd.to_datetime(df[KOL_REG], errors="coerce").dt.year
    return df


def je_firma(naziv, web) -> bool:
    n = "" if pd.isna(naziv) else str(naziv)
    if RE_FIRMA_DOSLOVNO.search(n) or RE_FIRMA_RIJEC.search(n):
        return True
    w = "" if pd.isna(web) else str(web).strip()
    return bool(w) and bool(RE_DOMEN.search(w))


def bod_oglasi(n: int) -> int:
    if n >= 1000:
        return 30
    if n >= 200:
        return 32
    if n >= 50:
        return 24
    if n >= 10:
        return 12
    return 0


def ocijeni(r: pd.Series) -> int:
    b = bod_oglasi(int(r[KOL_OGLASI]))
    b += {"Platinum": 25, "Gold": 12}.get(str(r[KOL_PAKET]), 5)
    if pd.notna(r[KOL_WEB]) and str(r[KOL_WEB]).strip():
        b += 15
    if r.get("tip_naloga") == "firma":
        b += 10
    god = r.get("godina_registracije")
    if pd.notna(god) and int(god) <= 2022:
        b += 8
    if str(r.get(KOL_PARTNER)).strip().upper() == "DA":
        b -= 10
    return int(b)


def kuka(r: pd.Series) -> str:
    """Kratak povod za prvi poziv, iz onoga sto se vidi u snimku."""
    n = int(r[KOL_OGLASI])
    paket = str(r[KOL_PAKET])
    if n == 0:
        return f"Placa {paket} paket a nema ni jedan aktivan oglas"
    if n >= 1000:
        return f"{n} oglasa, kvota obnova se sigurno ne trosi do kraja"
    if n >= 200:
        return f"{n} oglasa, rucna obnova je neizvediva bez alata"
    if n >= 50:
        return f"{n} oglasa, ima sta obnavljati i izdvajati"
    return f"{n} oglasa, prostor je u naslovima i kategorijama"


def pripremi(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["ocjena"] = df.apply(ocijeni, axis=1)
    df["procjena naknade KM/mj"] = df[KOL_PAKET].map(CIJENA).fillna(50).astype(int)
    df["povod za poziv"] = df.apply(kuka, axis=1)
    df["prioritet"] = pd.cut(df["ocjena"], bins=[-99, 39, 59, 999],
                             labels=["C", "B", "A"]).astype(str)
    return df


IZLAZNE = [KOL_SHOP, KOL_NAZIV, KOL_PAKET, KOL_OGLASI, KOL_GRAD, KOL_KANTON,
           KOL_WEB, "tip_naloga", "godina_registracije", "ocjena", "prioritet",
           "procjena naknade KM/mj", "povod za poziv", KOL_LINK]


def sortiraj(g: pd.DataFrame) -> pd.DataFrame:
    return g.sort_values(["ocjena", KOL_OGLASI], ascending=[False, False])


def formatiraj(dst: Path) -> None:
    zuto = PatternFill("solid", fgColor="FFF2CC")
    wb = load_workbook(dst)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        zaglavlje = [c.value for c in ws[1]]
        for c in ws[1]:
            c.font = Font(bold=True)
        if KOL_LINK in zaglavlje:
            i = zaglavlje.index(KOL_LINK) + 1
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=i)
                if isinstance(c.value, str) and c.value.startswith("http"):
                    c.hyperlink = c.value
                    c.font = Font(color="0563C1", underline="single")
        if KOL_PAKET in zaglavlje:
            i = zaglavlje.index(KOL_PAKET) + 1
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=i).value == "Platinum":
                    ws.cell(row=r, column=i).fill = zuto
        for i, naslov in enumerate(zaglavlje, start=1):
            duzine = [len(str(naslov))]
            for r in range(2, min(ws.max_row, 300) + 1):
                v = ws.cell(row=r, column=i).value
                if v is not None:
                    duzine.append(len(str(v)))
            ws.column_dimensions[get_column_letter(i)].width = min(max(duzine) + 2, 48)
    wb.save(dst)


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    argv = [a for a in sys.argv[1:] if not a.startswith("--")]
    for i, a in enumerate(sys.argv):
        if a == "--gold":
            CIJENA["Gold"] = int(sys.argv[i + 1])
        if a == "--platinum":
            CIJENA["Platinum"] = int(sys.argv[i + 1])
    src = Path(argv[0]).expanduser().resolve()
    if not src.exists():
        sys.exit(f"GRESKA: nema fajla {src}")
    m = re.search(r"(\d{4}-\d{2}-\d{2})", src.name)
    datum = m.group(1) if m else date.today().isoformat()
    dst = Path(argv[1]).expanduser().resolve() if len(argv) > 1 \
        else src.parent / f"prodajna-lista-{datum}.xlsx"

    df = pripremi(ucitaj(src))
    akt = df[df[KOL_OGLASI] > 0]
    prazni_plat = df[(df[KOL_OGLASI] == 0) & df[KOL_PAKET].eq("Platinum")]

    # Glavni bazen: aktivni sa najmanje 10 oglasa. Ispod toga nema sta optimizovati.
    bazen = akt[akt[KOL_OGLASI] >= 10]
    tiers = {t: sortiraj(bazen[bazen["prioritet"] == t]) for t in ("A", "B", "C")}
    mali = sortiraj(akt[akt[KOL_OGLASI] < 10])

    sazetak = pd.DataFrame({
        "grupa": ["A prioritet", "B prioritet", "C prioritet",
                  "Prazni Platinum", "Aktivni pod 10 oglasa", "UKUPNO lista"],
        "shopova": [len(tiers["A"]), len(tiers["B"]), len(tiers["C"]),
                    len(prazni_plat), len(mali),
                    len(bazen) + len(prazni_plat) + len(mali)],
        "Platinum": [int(t[KOL_PAKET].eq("Platinum").sum()) for t in tiers.values()]
                    + [len(prazni_plat), int(mali[KOL_PAKET].eq("Platinum").sum()),
                       int(pd.concat([bazen, prazni_plat, mali])[KOL_PAKET]
                           .eq("Platinum").sum())],
        "potencijal KM/mj": [int(t["procjena naknade KM/mj"].sum()) for t in tiers.values()]
                            + [int(prazni_plat["procjena naknade KM/mj"].sum()),
                               int(mali["procjena naknade KM/mj"].sum()),
                               int(pd.concat([bazen, prazni_plat, mali])
                                   ["procjena naknade KM/mj"].sum())],
        "ukupno oglasa": [int(t[KOL_OGLASI].sum()) for t in tiers.values()]
                         + [0, int(mali[KOL_OGLASI].sum()),
                            int(pd.concat([bazen, mali])[KOL_OGLASI].sum())],
    })

    with pd.ExcelWriter(dst, engine="openpyxl") as w:
        sazetak.to_excel(w, sheet_name="Sazetak", index=False)
        for t in ("A", "B", "C"):
            tiers[t][IZLAZNE].to_excel(w, sheet_name=f"{t} prioritet", index=False)
        sortiraj(prazni_plat)[IZLAZNE].to_excel(w, sheet_name="Prazni Platinum", index=False)
        mali[IZLAZNE].to_excel(w, sheet_name="Pod 10 oglasa", index=False)

    formatiraj(dst)
    print(f"Izvor: {src.name} (datum {datum})")
    print(sazetak.to_string(index=False))
    print(f"\nSnimljeno: {dst}")


if __name__ == "__main__":
    main()
