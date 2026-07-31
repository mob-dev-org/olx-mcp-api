"""Dopisuje telefon kandidata (Shop username) u postojeci Excel spisak.

OLX API ne vraca broj telefona ni za jedan tudji nalog (privatni podaci se ne vracaju za tudje
naloge), pa se telefon cita iz slobodnog teksta opisa shopa i oglasa preko CLI komande
`stats konkurent-telefon <username>` (regex pa Haiku, vidi src/core/telefon-ekstrakcija.ts).
Prinos zavisi od toga koliko kandidata je zaista upisalo broj u tekst, nije garantovan za sve.

Radi nad SVAKIM listom u ulaznom xlsx koji ima kolonu "Shop (username)" (prodajna-lista tier
listovi, razdvojeno po kantonima, ili sirovi snimak sa listom "SVI - zbirno"). Ostale listove
(sazetke, analize) prepisuje nepromijenjene. Isti username se ne pita API-ju dvaput ni kad se
pojavi na vise listova.

Ulazni fajl OSTAJE netaknut, izlaz je nov fajl (podrazumijevano <ulaz>-telefoni.xlsx).

Prije pokretanja mora postojati build: npm run build (dist/cli/index.js).

Upotreba:
    python3 dodaj-telefone.py <spisak.xlsx> [izlaz.xlsx] [--broj-oglasa N] [--pauza SEC]
        [--list "Ime lista,Drugi list"]

`--list` ogranicava API pozive na navedene listove (odvojeni zarezom, tacno ime lista); ostali
listovi se prepisuju u izlaz nepromijenjeni, bez novih kolona. Bez `--list` obradjuju se svi
listovi koji imaju kolonu "Shop (username)".
"""
import json
import subprocess
import sys
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

KOL_SHOP = "Shop (username)"
KOL_TELEFON = "Telefon"
KOL_IZVOR = "Telefon izvor"

REPO_ROOT = Path(__file__).resolve().parents[4]
DIST_CLI = REPO_ROOT / "dist" / "cli" / "index.js"


def telefon_za(username: str, broj_oglasa: int) -> dict:
    """Poziva CLI (stats konkurent-telefon) i vraca {telefon, izvor}. Ne baca gresku na
    pojedinacnom neuspjehu: jedan los username ne smije oboriti citav prolaz kroz spisak."""
    try:
        r = subprocess.run(
            ["node", str(DIST_CLI), "stats", "konkurent-telefon", username,
             "--broj-oglasa", str(broj_oglasa)],
            capture_output=True, text=True, timeout=60,
        )
        if r.returncode != 0:
            poruka = r.stderr.strip().splitlines()[-1] if r.stderr.strip() else "nepoznata greska"
            print(f"  {username}: GRESKA ({poruka})", file=sys.stderr)
            return {"telefon": None, "izvor": None}
        podaci = json.loads(r.stdout)
        return {"telefon": podaci.get("telefon"), "izvor": podaci.get("izvor")}
    except Exception as e:
        print(f"  {username}: GRESKA ({e})", file=sys.stderr)
        return {"telefon": None, "izvor": None}


def obradi_list(df: pd.DataFrame, broj_oglasa: int, pauza: float, cache: dict) -> pd.DataFrame:
    df = df.copy()
    telefoni, izvori = [], []
    for sirovi in df[KOL_SHOP]:
        username = str(sirovi).strip()
        if not username or username == "nan":
            telefoni.append(None)
            izvori.append(None)
            continue
        if username not in cache:
            cache[username] = telefon_za(username, broj_oglasa)
            time.sleep(pauza)
        telefoni.append(cache[username]["telefon"])
        izvori.append(cache[username]["izvor"])
    df[KOL_TELEFON] = telefoni
    df[KOL_IZVOR] = izvori
    return df


def formatiraj(dst: Path) -> None:
    wb = load_workbook(dst)
    for ws in wb.worksheets:
        zaglavlje = [c.value for c in ws[1]]
        if KOL_TELEFON not in zaglavlje:
            continue
        for c in ws[1]:
            c.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for i, naslov in enumerate(zaglavlje, start=1):
            duzine = [len(str(naslov))]
            for r in range(2, min(ws.max_row, 300) + 1):
                v = ws.cell(row=r, column=i).value
                if v is not None:
                    duzine.append(len(str(v)))
            ws.column_dimensions[get_column_letter(i)].width = min(max(duzine) + 2, 48)
    wb.save(dst)


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    if not DIST_CLI.exists():
        sys.exit(f"GRESKA: nema {DIST_CLI}. Prvo pokreni 'npm run build' u {REPO_ROOT}.")

    broj_oglasa = 5
    pauza = 0.4
    samo_listovi = None
    pozicioni = []
    argv = sys.argv[1:]
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == "--broj-oglasa":
            broj_oglasa = int(argv[i + 1])
            i += 2
            continue
        if a == "--pauza":
            pauza = float(argv[i + 1])
            i += 2
            continue
        if a == "--list":
            samo_listovi = {s.strip() for s in argv[i + 1].split(",")}
            i += 2
            continue
        pozicioni.append(a)
        i += 1

    src = Path(pozicioni[0]).expanduser().resolve()
    if not src.exists():
        sys.exit(f"GRESKA: nema fajla {src}")
    dst = Path(pozicioni[1]).expanduser().resolve() if len(pozicioni) > 1 \
        else src.with_name(f"{src.stem}-telefoni.xlsx")

    listovi = pd.ExcelFile(src, engine="openpyxl").sheet_names
    cache: dict = {}
    ukupno_sa_telefonom = 0
    ukupno_kandidata = 0

    with pd.ExcelWriter(dst, engine="openpyxl") as w:
        for naziv in listovi:
            df = pd.read_excel(src, sheet_name=naziv, engine="openpyxl")
            if KOL_SHOP not in df.columns or (samo_listovi is not None and naziv not in samo_listovi):
                df.to_excel(w, sheet_name=naziv, index=False)
                continue
            print(f"List '{naziv}': {len(df)} kandidata")
            df = obradi_list(df, broj_oglasa, pauza, cache)
            ukupno_kandidata += len(df)
            ukupno_sa_telefonom += int(df[KOL_TELEFON].notna().sum())
            df.to_excel(w, sheet_name=naziv, index=False)

    formatiraj(dst)

    regex_broj = sum(1 for v in cache.values() if v["izvor"] == "regex")
    haiku_broj = sum(1 for v in cache.values() if v["izvor"] == "haiku")
    print(f"\nGotovo: {dst}")
    print(f"Kandidata provjereno (jedinstvenih username-a): {len(cache)}")
    print(f"Telefon nadjen: {ukupno_sa_telefonom}/{ukupno_kandidata} redova "
          f"(regex {regex_broj}, Haiku {haiku_broj}, jedinstveno bez telefona "
          f"{len(cache) - regex_broj - haiku_broj})")


if __name__ == "__main__":
    main()
